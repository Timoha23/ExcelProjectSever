import re

import openpyxl
from openpyxl.cell import Cell
from openpyxl.styles import Alignment
from openpyxl.worksheet.merge import MergedCellRange
from openpyxl.worksheet.worksheet import Worksheet

from cutypes import FindCellTS, HeaderColumns, SensorData
from validators import HeaderDataModel


def find_cell_ts(excel_file: str, ts_number: str) -> FindCellTS:
    """
    Ищем координаты объединненой ячейки с номером ТС
    """

    wb = openpyxl.load_workbook(excel_file)

    result = {'error': 'ТС не найден'}

    # Флаг нашли ТС или нет
    cell_found = False

    # перебор страниц ексель книги
    for wsheet in wb.worksheets:
        # перебираем строки
        for cell in wsheet['A']:
            if cell.value is not None:
                if cell.value.lower() == ts_number.lower():
                    result = {'cell_ts': cell, 'wsheet': wsheet, 'wb': wb}
                    cell_found = True
                    merged_cells_starts = []
                    for merged_cell in wsheet.merged_cells:
                        merged_cells_starts.append(merged_cell.start_cell)
                    if cell not in merged_cells_starts:
                        return {
                            'error': ('Для корректной работы программы'
                                      ' необходимо самостоятельно '
                                      'заполнить два цикла')
                            }
        if cell_found:
            break
    return result


def find_header_in_sheet(wsheet: Worksheet) -> HeaderColumns:
    """
    Находим headedr файла wb страницы wsheet
    """
    global temperatures_length

    temperatures_length = 0
    target_header_data = {
        'Дата\nпроведения\nизмерений': 'date',
        'Цикл': 'cycle',
        'Фактическая\nглубина\nскважины, м':
        'actual_depth',
        'Глубина измерения, м': 'temperatures',
        'Высота надземной части скважины, м': 'height',
        'Глубина скв-ны с учётом надземной части, м': 'depth',
        't ср., ℃': 'avg_temp',
        'Температура\nокружающего\nвоздуха,℃': 'ambient_temperature',
    }

    header_col_dict = {}
    target_row = None

    # ищем строку с header
    for row in wsheet.rows:
        if row[0].value == 'Номер ТС':
            target_row = row
    if target_row:
        for col in target_row:
            if col.value in target_header_data.keys():
                header_col_dict[target_header_data[col.value]] = col
        try:
            HeaderDataModel(**header_col_dict)
        except Exception as ex:
            field_error = str(ex).split('\n')[1]
            return {'error': f'Ошибка. Поле хедера для {field_error} '
                             f'не найдено.'}
        return header_col_dict

    else:
        return {'error': 'Header не найден. В excel файле.'}


def get_header_lengths_and_input_row(
        wsheet: Worksheet,
        cell_ts: Cell,
        header_col_dict: HeaderColumns) -> tuple[int, bool]:
    """
    Находим строку окончание хедера, длину столбца с температурами и
    строку в которую вставляем данные
    """

    cycle_column_letter = header_col_dict['cycle'].column_letter
    cell_row = cell_ts.row
    if wsheet[f'{cycle_column_letter}{cell_row}'].value is None:
        input_row = cell_row
    else:
        input_row = cell_row + 1
    merged_cells = list(wsheet.merged_cells)
    for merged_cell in merged_cells:
        if cell_ts == merged_cell.start_cell:
            input_row = merged_cell.bottom[0][0] + 1
        elif header_col_dict['date'] == merged_cell.start_cell:
            after_row = merged_cell.max_row
        elif header_col_dict['temperatures'] == merged_cell.start_cell:
            temperatures_length = merged_cell.max_col - merged_cell.min_col + 1
    if input_row == cell_row:
        is_first_input = True
    else:
        is_first_input = False
    return (input_row, after_row, temperatures_length, is_first_input)


def unmerge_all_cells_after_header(
        wsheet: Worksheet,
        after_row: int) -> tuple[list[MergedCellRange], int]:
    """
    Разъединяем все ячейки снизу хедера
    """

    # переменная отвечающая за то, с какой строки разъединять,
    # 2 - дефолтная длина хедера за исключением первой строки

    merged_cells = list(wsheet.merged_cells)
    for merge in merged_cells:
        if merge.start_cell.row > after_row:
            wsheet.unmerge_cells(str(merge))
    return (merged_cells, after_row)


def add_new_row(wsheet: Worksheet, cell_ts: Cell, input_row: int) -> int:
    """
    Добавляем новую строку в excel файл
    """

    if cell_ts.row == input_row:
        pass
    else:
        wsheet.insert_rows(input_row)
    return input_row


def make_style_for_new_row(wsheet: Worksheet, input_row: int) -> None:
    """
    Переносим стили с верхней строки на текующую
    """

    for cell in wsheet[input_row]:
        cell_colum_letter = cell.column_letter
        cell_row = cell.row

        old_cell = wsheet[f'{cell_colum_letter}{cell_row-1}']
        if old_cell.has_style:
            cell._style = old_cell._style


def change_formuls_for_avg_temp(
        wsheet: Worksheet,
        input_row: int,
        header_in_cells: HeaderColumns) -> None:
    """
    Изменяем формулы для средней температуры
    """
    target_col = header_in_cells['avg_temp'].column_letter
    first_row = input_row + 1
    last_row = wsheet.max_row
    cell_range = wsheet[f"{target_col}{first_row}:{target_col}{last_row}"]
    found_cells = False
    for cell in cell_range:
        if cell[0].value is not None and '=AVERAGE' in str(cell[0].value):
            avg_cells: str = (cell[0].value.replace('=AVERAGE', '')
                              .replace('(', '').replace(')', '').split(':'))
            if found_cells is False:
                first_col = ''.join(x for x in avg_cells[0] if x.isalpha())
                last_col = ''.join(x for x in avg_cells[1] if x.isalpha())
                found_cells = True
            new_row = cell[0].row
            new_start_cell = first_col + str(new_row)
            new_end_cell = last_col + str(new_row)
            cell[0].value = f'=AVERAGE({new_start_cell}:{new_end_cell})'


def change_formuls_for_actual_depth(
        wsheet: Worksheet,
        input_row: int,
        header_in_cells: HeaderColumns) -> None:
    """
    Изменяем формулы для фактической глубины скважины
    """
    target_col = header_in_cells['actual_depth'].column_letter
    first_row = input_row + 1
    last_row = wsheet.max_row
    cell_range = wsheet[f"{target_col}{first_row}:{target_col}{last_row}"]
    depth_col = header_in_cells['depth'].column_letter
    height_col = header_in_cells['height'].column_letter
    for cell in cell_range:
        if cell[0].value is not None and '=' in str(cell[0].value):
            new_row = cell[0].row
            cell[0].value = f'=({depth_col}{new_row}-{height_col}{new_row})'


def make_merged_cells(
        wsheet: Worksheet,
        merged_cells: list[MergedCellRange],
        cell_ts: Cell,
        header_row: int,
        is_first_input: bool) -> None:
    """
    Соединяем ячейки обартно
    """

    # узнаем количество ячеек по мерджу
    const = 1
    if is_first_input:
        const = 0
    # объединяем ячейки
    for merged_cell in merged_cells:
        if merged_cell.start_cell.row == cell_ts.row:
            start_merge_column = merged_cell.top[0][1]
            start_merge_row = merged_cell.top[0][0]
            end_merge_column = merged_cell.bottom[0][1]
            end_merge_row = merged_cell.bottom[0][0] + 1
            wsheet.merge_cells(
                start_row=start_merge_row,
                start_column=start_merge_column,
                end_row=end_merge_row,
                end_column=end_merge_column
            )
        elif merged_cell.start_cell.row > cell_ts.row:
            start_merge_column = merged_cell.top[0][1]
            start_merge_row = merged_cell.top[0][0] + const
            end_merge_column = merged_cell.bottom[0][1]
            end_merge_row = merged_cell.bottom[0][0] + const
            wsheet.merge_cells(
                start_row=start_merge_row,
                start_column=start_merge_column,
                end_row=end_merge_row,
                end_column=end_merge_column
            )
        elif merged_cell.start_cell.row < cell_ts.row:
            wsheet.merge_cells(str(merged_cell))

        # текст по центру
        merged_cells = list(wsheet.merged_cells)
        for merged_cell in merged_cells:
            if merged_cell.start_cell.row > header_row:
                start_cell = merged_cell.start_cell
                start_cell.alignment = Alignment(vertical='center',
                                                 horizontal='center')


def put_data_to_excel(
        wsheet: Worksheet,
        input_row: int,
        header_in_cells: HeaderColumns,
        data: SensorData,
        temperatures_length: int,
        cell_ts: Cell) -> dict | None:
    """
    Вставляем данные в ексель файл
    """

    cargo_height = data['cargo_height']

    # имена колонок
    date_column = header_in_cells['date'].column_letter
    cycle_column = header_in_cells['cycle'].column_letter
    height_column = header_in_cells['height'].column_letter
    depth_column = header_in_cells['depth'].column_letter
    temperatures_column = header_in_cells['temperatures'].column
    actual_depth_column = header_in_cells['actual_depth'].column_letter
    ambient_temperature_column = (
        header_in_cells['ambient_temperature'].column_letter
    )
    avg_temperature_column = header_in_cells['avg_temp'].column_letter

    ambient_temperature = data['ambient_temperature']
    date = data['date']
    height = data['height'] + cargo_height
    depth = data['depth'] + cargo_height
    temperatures = data['temperatures']

    # вставляем дату
    cell_date = wsheet[f'{date_column}{input_row}']
    cell_date.value = date

    # вставляем цикл
    if cell_ts.row == input_row:
        cell_cycle = wsheet[f'{cycle_column}{input_row}']
        cell_cycle.value = '«0» цикл'
    else:
        old_cell_cycle = wsheet[f'{cycle_column}{input_row-1}']
        old_cell_cycle_num = old_cell_cycle.value
        cycle_number = (int(''.join(re.findall(r'[0-9]', old_cell_cycle_num)))
                        + 1)
        cell_cycle = wsheet[f'{cycle_column}{input_row}']
        cell_cycle.value = f'«{cycle_number}» цикл'

    # вставляем высоту
    old_cell_height = wsheet[f'{height_column}{input_row-1}']
    cell_height = wsheet[f'{height_column}{input_row}']
    cell_height.value = height
    if old_cell_height.has_style:
        cell_height._style = old_cell_height._style
    # вставляем глубину
    cell_depth = wsheet[f'{depth_column}{input_row}']
    cell_depth.value = depth

    # вставляем температуры
    # В случае когда температур больше необходимого, берем их с конца

    temp_column_counter = 0

    count_temps = depth - height
    if int(count_temps % 1 * 10) >= 9:
        count_temps = int(count_temps) + 1
    else:
        count_temps = int(count_temps)
    if len(temperatures) > count_temps:
        temperatures.reverse()
        temperatures = temperatures[:count_temps]
        temperatures.reverse()
    if len(temperatures) < temperatures_length:
        none_list = [None] * (temperatures_length - len(temperatures))
        temperatures.extend(none_list)

    # вставляем значения
    while temp_column_counter < temperatures_length:
        if temperatures[temp_column_counter] is None:
            wsheet.cell(
                input_row,
                temperatures_column + temp_column_counter
            ).value = '-'
        else:
            wsheet.cell(
                input_row,
                temperatures_column + temp_column_counter
            ).value = temperatures[temp_column_counter]
        temp_column_counter += 1

    # вставляем фактическую глубину
    cell_sum_depth = str(cell_depth.coordinate)
    cell_sum_height = str(cell_height.coordinate)
    wsheet[f'{actual_depth_column}{input_row}'] = (
        f'=({cell_sum_depth}-{cell_sum_height})'
    )

    # вставляем температуру окружающего воздуха
    cell_ambient_temperature = wsheet[f'{ambient_temperature_column}'
                                      f'{input_row}']
    cell_ambient_temperature.value = ambient_temperature

    # вставляем среднюю температуру
    prev_row = 1
    prev_avg_temp = wsheet[f'{avg_temperature_column}{input_row-prev_row}']
    while ('=СРЗНАЧ' not in str(prev_avg_temp.value) and
           '=AVERAGE' not in str(prev_avg_temp.value)):
        prev_row += 1
        if input_row - prev_row < 1:
            return {'error': 'Формула для среднего значения '
                             'температуры не найдена'}
        prev_avg_temp = wsheet[f'{avg_temperature_column}{input_row-prev_row}']
    avg_temp_column_for_formula = (
        prev_avg_temp.value.split('(')[1].replace(')', '').split(':')
    )
    first_col = (
        ''.join(x for x in avg_temp_column_for_formula[0] if x.isalpha())
    )
    last_col = (
        ''.join(x for x in avg_temp_column_for_formula[1] if x.isalpha())
    )
    cell_avg_temperature = wsheet[f'{avg_temperature_column}{input_row}']

    if type(wsheet[f'{first_col}{input_row}'].value) in (int, float):
        cell_avg_temperature.value = (
            f'=AVERAGE({first_col}{input_row}:{last_col}{input_row})'
        )
    else:
        cell_avg_temperature.value = '-'
