import re

import openpyxl
from openpyxl.cell import Cell
from openpyxl.styles import Alignment
from openpyxl.worksheet.merge import MergedCellRange
from openpyxl.worksheet.worksheet import Worksheet

from cutypes import FindCellTS, HeaderColumns, SensorData
from validators import HeaderDataModel


def find_cell_ts(excel_file: str, ts_number: str) -> FindCellTS:
    """
    Ищем координаты объединненой ячейки с номером ТС
    """

    wb = openpyxl.load_workbook(excel_file)

    result = {'error': 'ТС не найден'}

    # Флаг нашли ТС или нет
    cell_found = False

    # перебор страниц ексель книги
    for wsheet in wb.worksheets:
        # перебираем строки
        for cell in wsheet['A']:
            if cell.value is not None:
                if cell.value.lower() == ts_number.lower():
                    result = {'cell_ts': cell, 'wsheet': wsheet, 'wb': wb}
                    cell_found = True
                    merged_cells_starts = []
                    for merged_cell in wsheet.merged_cells:
                        merged_cells_starts.append(merged_cell.start_cell)
                    if cell not in merged_cells_starts:
                        return {
                            'error': ('Для корректной работы программы'
                                      ' необходимо самостоятельно '
                                      'заполнить два цикла')
                            }
        if cell_found:
            break
    return result


def find_header_in_sheet(wsheet: Worksheet) -> HeaderColumns:
    """
    Находим headedr файла wb страницы wsheet
    """
    global temperatures_length

    temperatures_length = 0
    target_header_data = {
        'Дата\nпроведения\nизмерений': 'date',
        'Цикл': 'cycle',
        'Фактическая\nглубина\nскважины, м':
        'actual_depth',
        'Глубина измерения, м': 'temperatures',
        'Высота надземной части скважины, м': 'height',
        'Глубина скв-ны с учётом надземной части, м': 'depth',
    }

    header_col_dict = {}
    target_row = None

    # ищем строку с header
    for row in wsheet.rows:
        if row[0].value == 'Номер ТС':
            target_row = row
    if target_row:
        for col in target_row:
            if col.value in target_header_data.keys():
                header_col_dict[target_header_data[col.value]] = col
        try:
            HeaderDataModel(**header_col_dict)
        except Exception as ex:
            field_error = str(ex).split('\n')[1]
            return {'error': f'Ошибка. Поле хедера для {field_error} '
                             f'не найдено.'}
        return header_col_dict

    else:
        return {'error': 'Header не найден. В excel файле.'}


def get_header_lengths_and_input_row(
        wsheet: Worksheet,
        cell_ts: Cell,
        header_col_dict: HeaderColumns) -> tuple[int, bool]:
    """
    Находим строку окончание хедера, длину столбца с температурами и
    строку в которую вставляем данные
    """

    cycle_column_letter = header_col_dict['cycle'].column_letter
    cell_row = cell_ts.row
    if wsheet[f'{cycle_column_letter}{cell_row}'].value is None:
        input_row = cell_row
    else:
        input_row = cell_row + 1
    merged_cells = list(wsheet.merged_cells)
    for merged_cell in merged_cells:
        if cell_ts == merged_cell.start_cell:
            input_row = merged_cell.bottom[0][0] + 1
        elif header_col_dict['date'] == merged_cell.start_cell:
            after_row = merged_cell.max_row
        elif header_col_dict['temperatures'] == merged_cell.start_cell:
            temperatures_length = merged_cell.max_col - merged_cell.min_col + 1
    if input_row == cell_row:
        is_first_input = True
    else:
        is_first_input = False
    return (input_row, after_row, temperatures_length, is_first_input)


def unmerge_all_cells_after_header(
        wsheet: Worksheet,
        after_row: int) -> tuple[list[MergedCellRange], int]:
    """
    Разъединяем все ячейки снизу хедера
    """

    # переменная отвечающая за то, с какой строки разъединять,
    # 2 - дефолтная длина хедера за исключением первой строки

    merged_cells = list(wsheet.merged_cells)
    for merge in merged_cells:
        if merge.start_cell.row > after_row:
            wsheet.unmerge_cells(str(merge))
    return (merged_cells, after_row)


def add_new_row(wsheet: Worksheet, cell_ts: Cell, input_row: int) -> int:
    """
    Добавляем новую строку в excel файл
    """

    if cell_ts.row == input_row:
        pass
    else:
        wsheet.insert_rows(input_row)
    return input_row


def make_style_for_new_row(wsheet: Worksheet, input_row: int) -> None:
    """
    Переносим стили с верхней строки на текующую
    """

    for cell in wsheet[input_row]:
        cell_colum_letter = cell.column_letter
        cell_row = cell.row

        old_cell = wsheet[f'{cell_colum_letter}{cell_row-1}']
        if old_cell.has_style:
            cell._style = old_cell._style


def make_merged_cells(
        wsheet: Worksheet,
        merged_cells: list[MergedCellRange],
        cell_ts: Cell,
        header_row: int,
        is_first_input: bool) -> None:
    """
    Соединяем ячейки обартно
    """

    # узнаем количество ячеек по мерджу
    const = 1
    if is_first_input:
        const = 0
    # объединяем ячейки
    for merged_cell in merged_cells:
        if merged_cell.start_cell.row == cell_ts.row:
            start_merge_column = merged_cell.top[0][1]
            start_merge_row = merged_cell.top[0][0]
            end_merge_column = merged_cell.bottom[0][1]
            end_merge_row = merged_cell.bottom[0][0] + 1
            wsheet.merge_cells(
                start_row=start_merge_row,
                start_column=start_merge_column,
                end_row=end_merge_row,
                end_column=end_merge_column
            )
        elif merged_cell.start_cell.row > cell_ts.row:
            start_merge_column = merged_cell.top[0][1]
            start_merge_row = merged_cell.top[0][0] + const
            end_merge_column = merged_cell.bottom[0][1]
            end_merge_row = merged_cell.bottom[0][0] + const
            wsheet.merge_cells(
                start_row=start_merge_row,
                start_column=start_merge_column,
                end_row=end_merge_row,
                end_column=end_merge_column
            )
        elif merged_cell.start_cell.row < cell_ts.row:
            wsheet.merge_cells(str(merged_cell))

        # текст по центру
        merged_cells = list(wsheet.merged_cells)
        for merged_cell in merged_cells:
            if merged_cell.start_cell.row > header_row:
                start_cell = merged_cell.start_cell
                start_cell.alignment = Alignment(vertical='center',
                                                 horizontal='center')


def put_data_to_excel(
        wsheet: Worksheet,
        input_row: int,
        header_in_cells: HeaderColumns,
        data: SensorData,
        temperatures_length: int,
        cell_ts: Cell) -> dict | None:
    """
    Вставляем данные в ексель файл
    """

    cargo_height = data['cargo_height']

    # имена колонок
    date_column = header_in_cells['date'].column_letter
    cycle_column = header_in_cells['cycle'].column_letter
    height_column = header_in_cells['height'].column_letter
    depth_column = header_in_cells['depth'].column_letter
    temperatures_column = header_in_cells['temperatures'].column
    actual_depth_column = header_in_cells['actual_depth'].column_letter

    date = data['date']
    height = data['height'] + cargo_height
    depth = data['depth'] + cargo_height
    temperatures = data['temperatures']

    # вставляем дату
    old_cell_date = wsheet[f'{date_column}{input_row-1}']
    cell_date = wsheet[f'{date_column}{input_row}']
    # put_date = put_date.strftime("%m/%d/%Y")
    cell_date.value = date
    if old_cell_date.has_style:
        cell_date._style = old_cell_date._style

    # вставляем цикл
    if cell_ts.row == input_row:
        cell_cycle = wsheet[f'{cycle_column}{input_row}']
        cell_cycle.value = '«0» цикл'
        # cell_cycle._style = StyleArray('i', [3, 0, 3, 0, 0, 3, 0, 0, 0])
    else:
        old_cell_cycle = wsheet[f'{cycle_column}{input_row-1}']
        old_cell_cycle_num = old_cell_cycle.value
        cycle_number = (int(''.join(re.findall(r'[0-9]', old_cell_cycle_num)))
                        + 1)
        cell_cycle = wsheet[f'{cycle_column}{input_row}']
        cell_cycle.value = f'«{cycle_number}» цикл'
        if old_cell_cycle.has_style:
            cell_cycle._style = old_cell_cycle._style

    # вставляем высоту
    old_cell_height = wsheet[f'{height_column}{input_row-1}']
    cell_height = wsheet[f'{height_column}{input_row}']
    cell_height.value = height
    if old_cell_height.has_style:
        cell_height._style = old_cell_height._style
    # вставляем глубину
    old_cell_depth = wsheet[f'{depth_column}{input_row-1}']
    cell_depth = wsheet[f'{depth_column}{input_row}']
    cell_depth.value = depth
    if old_cell_depth.has_style:
        cell_depth._style = old_cell_depth._style

    # вставляем температуры
    # В случае когда температур больше необходимого, берем их с конца

    temp_column_counter = 0

    count_temps = depth - height
    if int(count_temps % 1 * 10) >= 9:
        count_temps = int(count_temps) + 1
    else:
        count_temps = int(count_temps)
    if len(temperatures) > count_temps:
        temperatures.reverse()
        temperatures = temperatures[:count_temps]
        temperatures.reverse()
    if len(temperatures) < temperatures_length:
        none_list = [None] * (temperatures_length - len(temperatures))
        temperatures.extend(none_list)

    # вставляем значения
    while temp_column_counter < temperatures_length:
        if temperatures[temp_column_counter] is None:
            wsheet.cell(
                input_row,
                temperatures_column + temp_column_counter
            ).value = '-'
        else:
            wsheet.cell(
                input_row,
                temperatures_column + temp_column_counter
            ).value = temperatures[temp_column_counter]
        temp_column_counter += 1

    # вставляем фактическую глубину
    cell_sum_depth = str(cell_depth.coordinate)
    cell_sum_height = str(cell_height.coordinate)
    wsheet[f'{actual_depth_column}{input_row}'] = (
        f'=({cell_sum_depth}-{cell_sum_height})'
    )
