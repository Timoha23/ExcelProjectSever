import openpyxl
from openpyxl.utils.exceptions import InvalidFileException
from pydantic.error_wrappers import ValidationError

from cutypes import SensorData
from validators import SensorDataModel


def get_sensor_data(number: str, sensor_data_path: str) -> SensorData:
    """
    Получаем данные из файла ПКЦД.xlsx
    Данная функция принимает number (n734)
    """

    path = sensor_data_path

    try:
        wb = openpyxl.load_workbook(path)
    except (FileNotFoundError, InvalidFileException):
        return {'error': f'Неверно указан путь с датчиками: {path}.'
                         f' Проверьте файл settings.txt'}

    ws = wb.active
    rows = ws.max_row
    data = {'temperatures': []}

    for i in range(1, rows+1):
        if ws[f'A{i}'].value == number:
            sensors_count = int(ws[i+1][2].value.split(':')[1])
            data['date'] = ws[i][1].value  # дата
            data['gk_name'] = ws[i+1][4+sensors_count].value  # имя ГК
            data['ts_number'] = ws[i+1][5+sensors_count].value  # номер ТС
            data['depth'] = ws[i+1][6+sensors_count].value  # глубина
            data['height'] = ws[i+1][7+sensors_count].value  # высота
            # высота груза
            data['cargo_height'] = ws[i+1][8+sensors_count].value

            for index, el in enumerate(ws[i+1][4:]):
                if index < sensors_count:
                    data['temperatures'].append(el.value)
            try:
                SensorDataModel(**data)
            except ValidationError as ex:
                field_error = str(ex).split('\n')[1]
                return {'error': f'Ошибка в поле {field_error}.'
                                 f' Проверьте данное поле'}
            return data
    return {'error': f'Номер датчика {number} не найден.'}
